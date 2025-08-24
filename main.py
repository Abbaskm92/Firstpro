import argparse
import re
import os
import time
import math
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill

# ================= Helpers =================
def safe_sheet_name(name: str) -> str:
    """Excel sheet name max 31 chars; remove illegal chars."""
    return re.sub(r'[:\\/?*\[\]]', "_", str(name))[:31]

def to_number(x):
    """Turn '1,200' or ' 300 ' into float; return None if empty."""
    if pd.isna(x) or x == "":
        return None
    s = str(x).replace(",", "").strip()
    try:
        return float(s)
    except Exception:
        return None

def normalize_dates(series: pd.Series) -> pd.Series:
    """
    Parse timestamps robustly. If tz-aware, convert to Asia/Baghdad then drop tz (naive).
    If tz-naive, just parse and keep naive.
    """
    dt = pd.to_datetime(series, errors="coerce", utc=True)
    if getattr(dt.dt, "tz", None) is not None:
        return dt.dt.tz_convert("Asia/Baghdad").dt.tz_localize(None)
    dt2 = pd.to_datetime(series, errors="coerce")
    if getattr(dt2.dt, "tz", None) is not None:
        return dt2.dt.tz_convert("Asia/Baghdad").dt.tz_localize(None)
    return dt2

def ensure_row_merge(ws, row, c1, c2):
    """
    On a single row, remove any merge overlapping columns [c1..c2] and then
    merge exactly (row, c1) .. (row, c2).
    """
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row:
            if not (rng.max_col < c1 or rng.min_col > c2):
                to_unmerge.append(str(rng))
    for coord in to_unmerge:
        try:
            ws.unmerge_cells(coord)
        except Exception:
            pass
    try:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    except Exception:
        pass  # avoid Excel repairs if any odd objects exist

def unmerge_covering(ws, row, col_idx):
    """Unmerge any merged range that covers cell (row, col_idx)."""
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col_idx <= rng.max_col:
            to_unmerge.append(str(rng))
    for coord in to_unmerge:
        try:
            ws.unmerge_cells(coord)
        except Exception:
            pass

def clone_style(dst, src, fallback_num_fmt=None):
    """Copy core style attributes from src to dst. Skip on merged non-anchors."""
    if isinstance(dst, MergedCell):
        return
    try:
        dst.font = src.font
        dst.border = src.border
        dst.fill = src.fill
        dst.number_format = fallback_num_fmt if fallback_num_fmt else src.number_format
        dst.protection = src.protection
        dst.alignment = src.alignment
    except Exception:
        pass

def get_cell(ws, row, letter):
    return ws[f"{letter}{row}"]

def style_prototypes(ws, item_row, name_col="B", price_col="E", comm_col="F", final_col="G"):
    """Capture style prototypes from the template's first item row."""
    proto = {}
    proto["name"] = get_cell(ws, item_row, name_col)           # B (name anchor)
    proto["price"] = get_cell(ws, item_row, price_col)         # E
    proto["comm"]  = get_cell(ws, item_row, comm_col)          # F
    proto["final"] = get_cell(ws, item_row, final_col)         # G
    proto["num_fmt"] = proto["price"].number_format or "#,##0.00"
    return proto

# ---- robust save (handles file open/locked) ----
def safe_save_workbook(wb, out_path: Path, mode: str = "unique") -> Path:
    """
    Save workbook safely.
      mode='unique'     -> create unique path if target blocked (… (1).xlsx, (2)…)
      mode='overwrite'  -> save to temp then atomic replace; fallback to unique on PermissionError
      mode='timestamp'  -> add _YYYYMMDD_HHMMSS suffix
    Returns the actual Path saved to.
    """
    out_path = Path(out_path)
    suffix = out_path.suffix or ".xlsx"

    if mode == "timestamp":
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        stamped = out_path.with_name(f"{out_path.stem}_{ts}{suffix}")
        wb.save(stamped)
        return stamped

    if mode == "overwrite":
        tmp = out_path.with_name(f"{out_path.stem}.tmp_{int(time.time())}{suffix}")
        wb.save(tmp)
        try:
            os.replace(tmp, out_path)  # atomic on Windows/Unix
            return out_path
        except PermissionError:
            pass
        try:
            if tmp.exists():
                tmp.unlink()
        except Exception:
            pass
        mode = "unique"

    # unique mode
    candidate = out_path
    i = 1
    while True:
        try:
            wb.save(candidate)
            return candidate
        except PermissionError:
            candidate = out_path.with_name(f"{out_path.stem} ({i}){suffix}")
            i += 1

# ---- wrapping & row-height helpers ----
def get_col_width(ws, col_letter, default_width=8.43):
    """Return effective column width; if not set in sheet, return Excel default."""
    w = ws.column_dimensions[col_letter].width
    return float(w) if w else default_width

def estimate_needed_lines(text: str, total_width_chars: float, min_chars_per_line: int = 10) -> int:
    """
    Roughly estimate how many wrapped lines are needed based on merged width (in 'Excel width units').
    """
    if not text:
        return 1
    parts = str(text).splitlines() or [str(text)]
    lines = 0
    chars_per_line = max(min_chars_per_line, int(total_width_chars * 1.1))
    for seg in parts:
        seg = seg.strip()
        if not seg:
            lines += 1
            continue
        lines += max(1, math.ceil(len(seg) / chars_per_line))
    return lines

def apply_wrap_and_autoheight(ws, row, merged_cols=("B","C","D"), base_height=15, min_height=15, max_height=180):
    """
    Enable wrapping in B{row} (anchor of merged B:D) and increase row height to fit.
    """
    total_width = sum(get_col_width(ws, c) for c in merged_cols)
    cell = ws[f"{merged_cols[0]}{row}"]
    text = "" if cell.value is None else str(cell.value)
    needed_lines = estimate_needed_lines(text, total_width)
    new_height = max(min_height, min(max_height, base_height * needed_lines))
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c in merged_cols[1:]:
        mc = ws[f"{c}{row}"]
        if not isinstance(mc, MergedCell):
            mc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = new_height

# ---- white background helper ----
def apply_white_background(ws, start_row, end_row, cols=("B","C","D","E","F","G")):
    """Fill the specified range with white background."""
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for row in range(start_row, end_row + 1):
        for col in cols:
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MergedCell):
                cell.fill = white_fill

# ================= Main work =================
def fill_invoice_per_vendor(
    df: pd.DataFrame,
    template_file: Path,
    out_file: Path,
    month_str: str,
    vendor_col="Business Name",
    item_col="Qty & Product Name",
    price_col="Product Price",
    commission_col="Commission Amount",   # accepts 'Commision Amount' via fallback
    date_col="Created time",
    vendor_cell="B12",
    invoice_code_cell="F12",   # merged F12:G12 (write to F12)
    invoice_date_cell="F15",   # merged F15:G15 (write to F15)
    header_date_cell="B9",     # also write today's date here
    first_item_row=19,
    name_merge_cols=("B", "D"),   # merge B..D for item name
    price_col_letter="E",
    commission_col_letter="F",
    final_col_letter="G",
    base_sum_row=21,    # when 1 item -> G21
    base_total_row=23,  # when 1 item -> F:G23 (we write to G23)
    out_mode="unique",
):
    # Validate inputs
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_file}")
    if df.empty:
        raise ValueError("Input DataFrame is empty")

    # Ensure required columns exist (exact spellings, with commission typo fallback)
    required_cols = [vendor_col, item_col, price_col, commission_col, date_col]
    missing = [c for c in required_cols if c not in df.columns]
    if "Commission Amount" in missing and "Commision Amount" in df.columns:
        commission_col = "Commision Amount"
        required_cols = [vendor_col, item_col, price_col, commission_col, date_col]
        missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    # Normalize dates to naive Baghdad local time
    df = df.copy()
    df[date_col] = normalize_dates(df[date_col])

    # Parse month window
    try:
        year = int(month_str.split("-")[0])
        month = int(month_str.split("-")[1])
        if month < 1 or month > 12:
            raise ValueError
    except Exception:
        raise ValueError("Month format must be YYYY-MM (e.g., 2025-07)")
    start = pd.Timestamp(year, month, 1)
    end = start + pd.offsets.MonthBegin(1)

    # Filter to desired month
    dfm = df[(df[date_col] >= start) & (df[date_col] < end)]

    # Load template
    wb = load_workbook(template_file)
    template_ws = wb.worksheets[0]

    # Vendors for this month
    vendors = sorted(v for v in dfm[vendor_col].dropna().astype(str).unique())

    # Today strings for invoice code and dates
    today_str = datetime.now().strftime("%Y%m%d")  # e.g., 20250724
    today_date = datetime.now().date()

    # Column indices
    start_col_idx = column_index_from_string(name_merge_cols[0])   # B
    end_col_idx = column_index_from_string(name_merge_cols[1])     # D
    e_idx = column_index_from_string(price_col_letter)             # E
    f_idx = column_index_from_string(commission_col_letter)        # F
    g_idx = column_index_from_string(final_col_letter)             # G

    # Prepare style prototypes from the template's first item row
    proto = style_prototypes(
        template_ws,
        first_item_row,
        name_col=name_merge_cols[0],
        price_col=price_col_letter,
        comm_col=commission_col_letter,
        final_col=final_col_letter,
    )
    num_fmt = proto["num_fmt"] if proto.get("num_fmt") else "#,##0.00"

    made_any = 0

    for idx, vendor in enumerate(vendors, start=1):
        sub = dfm[dfm[vendor_col].astype(str) == vendor]
        if sub.empty:
            continue

        # Copy the template per vendor
        ws = wb.copy_worksheet(template_ws)
        ws.title = safe_sheet_name(vendor)

        # HIDE GRIDLINES for this vendor sheet (view + print)
        ws.sheet_view.showGridLines = False
        ws.print_options.gridLines = False

        # Header fields
        ws[vendor_cell] = vendor
        ws[invoice_code_cell] = f"{today_str}-{idx}"  # Sequential invoice code: YYYYMMDD-<n>
        ws[invoice_date_cell] = today_date
        ws[header_date_cell] = today_date

        # Number of items for this vendor in the month
        n_items = len(sub)
        if n_items == 0:
            continue

        # Add extra rows so items start at first_item_row and extend downward
        if n_items > 1:
            ws.insert_rows(first_item_row + 1, amount=n_items - 1)

        # Ensure the item rows are visible (avoid inheriting hidden row formatting)
        template_height = ws.row_dimensions[first_item_row].height
        for i in range(n_items):
            r = first_item_row + i
            ws.row_dimensions[r].hidden = False
            if template_height is not None:
                ws.row_dimensions[r].height = template_height

        # --- Apply uniform styles BEFORE merging ---
        for i in range(n_items):
            r = first_item_row + i
            # Name block B,C,D — style each cell first, then merge later
            for col_letter in ("B", "C", "D"):
                dcell = get_cell(ws, r, col_letter)
                clone_style(dcell, proto["name"])
            # Numeric columns E/F/G
            e_cell = get_cell(ws, r, price_col_letter)
            f_cell = get_cell(ws, r, commission_col_letter)
            g_cell = get_cell(ws, r, final_col_letter)
            clone_style(e_cell, proto["price"], fallback_num_fmt=num_fmt)
            clone_style(f_cell, proto["comm"], fallback_num_fmt=num_fmt)
            clone_style(g_cell, proto["final"], fallback_num_fmt=num_fmt)

        # Merge B..D for each item row safely (remove overlaps first)
        for i in range(n_items):
            r = first_item_row + i
            ensure_row_merge(ws, r, start_col_idx, end_col_idx)

        # Write item rows (ensure E/F/G are NOT merged for this row)
        sub_rows = sub[[item_col, price_col, commission_col]].fillna("")
        for i, (_, row) in enumerate(sub_rows.iterrows()):
            r = first_item_row + i
            # make sure columns E, F, G for this row aren't part of any merge
            unmerge_covering(ws, r, e_idx)
            unmerge_covering(ws, r, f_idx)
            unmerge_covering(ws, r, g_idx)

            # Write item text into B (anchor of the merged B:D), enable wrap+autoheight
            item_text = row[item_col]
            ws[f"B{r}"] = item_text
            apply_wrap_and_autoheight(ws, r, merged_cols=("B", "C", "D"))

            # Numeric values
            price_val = to_number(row[price_col])
            comm_val = to_number(row[commission_col])
            ws[f"{price_col_letter}{r}"] = price_val
            ws[f"{commission_col_letter}{r}"] = comm_val
            ws[f"{final_col_letter}{r}"] = (
                price_val - comm_val
                if (price_val is not None and comm_val is not None)
                else None
            )
            # enforce number format
            ws[f"{price_col_letter}{r}"].number_format = num_fmt
            ws[f"{commission_col_letter}{r}"].number_format = num_fmt
            ws[f"{final_col_letter}{r}"].number_format = num_fmt

        # Clear the immediate spacer row after the last item (fix stray PRODUCT formula)
        spacer_row = first_item_row + n_items
        for col in ("E", "F", "G"):
            ws[f"{col}{spacer_row}"] = None

        # Sum row and total row, adjusted by item count
        sum_row = (base_sum_row - 1) + n_items      # 1 item -> 21, 2 items -> 22, etc.
        total_row = (base_total_row - 1) + n_items  # 1 item -> 23, 2 items -> 24, etc.
        start_sum_row = first_item_row
        end_sum_row = first_item_row + n_items - 1

        # Unmerge any F/G merges covering summary rows so G is writeable/visible
        for rr in (sum_row, total_row):
            unmerge_covering(ws, rr, e_idx)
            unmerge_covering(ws, rr, f_idx)
            unmerge_covering(ws, rr, g_idx)

        # Place SUM of finals into G{sum_row}
        ws[f"{final_col_letter}{sum_row}"] = f"=SUM({final_col_letter}{start_sum_row}:{final_col_letter}{end_sum_row})"
        ws[f"{final_col_letter}{sum_row}"].number_format = num_fmt

        # Total row: same total in G{total_row}, and ensure F{total_row} is blank
        ws[f"{final_col_letter}{total_row}"] = f"={final_col_letter}{sum_row}"
        ws[f"{final_col_letter}{total_row}"].number_format = num_fmt
        ws[f"{commission_col_letter}{total_row}"] = None

        # Make G{total_row} bold and size 16 (visual emphasis)
        ws[f"{final_col_letter}{total_row}"].font = Font(bold=True, size=16)

        # Ensure summary rows visible & styled uniformly
        ws.row_dimensions[sum_row].hidden = False
        ws.row_dimensions[total_row].hidden = False
        for col_letter in (price_col_letter, commission_col_letter, final_col_letter):
            clone_style(get_cell(ws, sum_row, col_letter), proto["final"], fallback_num_fmt=num_fmt)
            clone_style(get_cell(ws, total_row, col_letter), proto["final"], fallback_num_fmt=num_fmt)

        # Force white background from first item row through total row (B..G)
        apply_white_background(ws, first_item_row, total_row, cols=("B", "C", "D", "E", "F", "G"))

        made_any += 1

    # Remove the original template sheet so only vendor sheets remain
    if template_ws.title in wb.sheetnames:
        try:
            wb.remove(template_ws)
        except Exception:
            pass

    # Final sweep: hide gridlines on ALL sheets (view + print)
    for sname in wb.sheetnames:
        ws2 = wb[sname]
        ws2.sheet_view.showGridLines = False
        ws2.print_options.gridLines = False

    # --- Save safely ---
    actual_path = safe_save_workbook(wb, Path(out_file), mode=out_mode)
    print(f"Invoice sheets created: {made_any}")
    print(f"Saved to: {actual_path}")

# ================= CLI / Entry =================
if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--file", required=True, help="Path to Test Sample.xlsx")
    p.add_argument("--template", required=True, help="Path to Invoice test.xlsx")
    p.add_argument("--out", default="Invoice_Output.xlsx", help="Output file name")
    p.add_argument(
        "--out-mode",
        choices=["unique", "overwrite", "timestamp"],
        default="unique",
        help="How to handle existing/locked output files",
    )
    args = p.parse_args()

    # Interactive prompts for year and month
    year_input = input("Enter the year (e.g., 2025): ").strip()
    month_input = input("Enter the month number (1-12): ").strip()

    try:
        year = int(year_input)
        month = int(month_input)
        if month < 1 or month > 12:
            raise ValueError
    except ValueError:
        raise ValueError("Invalid input. Please enter a valid year and a month between 1 and 12.")

    # Format month as YYYY-MM (zero-padded month)
    month_str = f"{year}-{month:02d}"

    df = pd.read_excel(args.file)
    fill_invoice_per_vendor(
        df=df,
        template_file=Path(args.template),
        out_file=Path(args.out),
        month_str=month_str,
        out_mode=args.out_mode,
    )